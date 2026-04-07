"""
小红书笔记采集器 - Flask 服务器
运行: python app.py
访问: http://localhost:5000
"""

import os
import json
import threading
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_file, Response, stream_with_context
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from scraper import scrape_note, do_login, check_login_from_storage, parse_share_text

app = Flask(__name__)

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
EXCEL_PATH = DATA_DIR / "xhs_notes.xlsx"

HEADERS = ["序号", "标题", "作者", "发布日期", "正文内容", "Hook类型", "话题标签", "点赞数", "收藏数", "评论数", "原链接", "笔记链接", "采集时间"]

# 登录状态：idle | logging_in | done | failed
_login_status = {"state": "idle", "message": ""}
_login_lock = threading.Lock()


# ── Excel 工具 ──────────────────────────────────────────────────

def _ensure_excel():
    """确保 Excel 文件存在并有表头"""
    if EXCEL_PATH.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "小红书笔记"
    ws.append(HEADERS)
    _style_header(ws)
    wb.save(EXCEL_PATH)


def _style_header(ws):
    fill = PatternFill("solid", fgColor="FF2442")  # 小红书红
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 列宽
    widths = [6, 40, 16, 12, 50, 12, 50, 8, 8, 8, 35, 45, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22


def _append_to_excel(row_data: dict) -> int:
    """追加一行到 Excel，返回当前总行数（不含表头）"""
    _ensure_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    next_row = ws.max_row + 1
    seq = next_row - 1  # 序号（第1行是表头）
    ws.append([
        seq,
        row_data.get("title", ""),
        row_data.get("author", ""),
        row_data.get("date", ""),
        row_data.get("content", ""),
        row_data.get("hook_type", ""),
        row_data.get("tags", ""),
        row_data.get("likes", ""),
        row_data.get("collects", ""),
        row_data.get("comments", ""),
        row_data.get("original_url", ""),
        row_data.get("final_url") or row_data.get("original_url", ""),
        row_data.get("scraped_at", ""),
    ])
    # 正文内容列自动换行
    for cell in ws[next_row]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(EXCEL_PATH)
    return seq


def _find_duplicate(url: str) -> dict | None:
    """检查 URL 是否已存在于 Excel，返回已有行（字典）或 None"""
    if not url or not EXCEL_PATH.exists():
        return None
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    url = url.strip().rstrip("/")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        d = dict(zip(HEADERS, row))
        for field in ("原链接", "笔记链接"):
            existing = str(d.get(field) or "").strip().rstrip("/")
            if existing and existing == url:
                return d
    return None


def _read_all_from_excel() -> list:
    """读取 Excel 全部数据（不含表头），返回字典列表"""
    if not EXCEL_PATH.exists():
        return []
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(dict(zip(HEADERS, row)))
    return rows


# 脏行判断关键词
_DIRTY_TITLES   = {"手机号登录", "验证码登录", "登录", "小红书"}
_DIRTY_CONTENTS = {"", "（未登录，无正文）", None}


def _is_dirty(row: dict) -> bool:
    """判断一行是否为脏数据"""
    title   = str(row.get("标题") or "").strip()
    content = str(row.get("正文内容") or "").strip()
    url     = str(row.get("原链接") or row.get("笔记链接") or "").strip()
    if not url:
        return False  # 没有链接，无法修复，跳过
    return title in _DIRTY_TITLES or content in _DIRTY_CONTENTS


def _update_excel_row(seq: int, row_data: dict):
    """按序号原地更新 Excel 中对应行"""
    if not EXCEL_PATH.exists():
        return
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    # 列索引映射（1-based）
    col_map = {h: i + 1 for i, h in enumerate(HEADERS)}

    for excel_row in ws.iter_rows(min_row=2):
        if excel_row[0].value == seq:  # 第1列是序号
            def _set(header, value):
                if value:
                    c = excel_row[col_map[header] - 1]
                    c.value = value
                    c.alignment = Alignment(wrap_text=True, vertical="top")

            _set("标题",    row_data.get("title"))
            _set("作者",    row_data.get("author"))
            _set("发布日期", row_data.get("date"))
            _set("正文内容", row_data.get("content"))
            _set("Hook类型", row_data.get("hook_type"))
            _set("话题标签", row_data.get("tags"))
            _set("点赞数",  row_data.get("likes"))
            _set("收藏数",  row_data.get("collects"))
            _set("评论数",  row_data.get("comments"))
            _set("笔记链接", row_data.get("final_url") or row_data.get("original_url"))
            _set("采集时间", row_data.get("scraped_at"))

            # 修复行改为绿色底色标记
            green = PatternFill("solid", fgColor="E8F5E9")
            for cell in excel_row:
                if cell.fill.fgColor.rgb in ("00000000", "FF000000", ""):
                    cell.fill = green
            break

    wb.save(EXCEL_PATH)


# ── 路由 ─────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/login-status")
def login_status():
    with _login_lock:
        state = _login_status.copy()
    state["logged_in"] = check_login_from_storage()
    return jsonify(state)


@app.route("/login", methods=["POST"])
def login():
    with _login_lock:
        if _login_status["state"] == "logging_in":
            return jsonify({"ok": False, "message": "登录流程正在进行中，请在浏览器窗口扫码"})
        _login_status["state"] = "logging_in"
        _login_status["message"] = "浏览器已打开，请用小红书 App 扫描二维码"

    def run_login():
        success = do_login()
        with _login_lock:
            if success:
                _login_status["state"] = "done"
                _login_status["message"] = "登录成功！"
            else:
                _login_status["state"] = "failed"
                _login_status["message"] = "登录超时或失败，请重试"

    t = threading.Thread(target=run_login, daemon=True)
    t.start()
    return jsonify({"ok": True, "message": "浏览器已打开，请扫码登录"})


@app.route("/parse", methods=["POST"])
def parse():
    """预解析分享文本，返回提取到的 url/title/author，供前端预览"""
    body = request.get_json(silent=True) or {}
    text = (body.get("text") or "").strip()
    return jsonify(parse_share_text(text))


@app.route("/scrape", methods=["POST"])
def scrape():
    body = request.get_json(silent=True) or {}
    # 支持直接传 raw 分享文本（text 字段）或纯链接（url 字段）
    raw_text = (body.get("text") or body.get("url") or "").strip()
    if not raw_text:
        return jsonify({"success": False, "error": "请提供链接或分享文本"}), 400

    parsed = parse_share_text(raw_text)
    url = parsed.get("url") or ""
    if not url:
        return jsonify({"success": False, "error": "未能从输入中提取到有效链接"}), 400

    force = bool(body.get("force", False))

    # 未 force 时先查重
    if not force:
        dup = _find_duplicate(url)
        if dup:
            return jsonify({
                "success": False,
                "duplicate": True,
                "existing": {
                    "seq":   dup.get("序号"),
                    "title": dup.get("标题", ""),
                    "author": dup.get("作者", ""),
                    "date":  dup.get("发布日期", ""),
                    "scraped_at": dup.get("采集时间", ""),
                },
                "url": url,
            })

    result = scrape_note(url)
    if result["success"]:
        # 用分享文本中解析的标题/作者作为兜底（网页采集优先）
        data = result["data"]
        if not data.get("title") and parsed.get("title"):
            data["title"] = parsed["title"]
        if not data.get("author") and parsed.get("author"):
            data["author"] = parsed["author"]
        # 保留用户粘贴的原始文本作为 original_url
        if not data.get("original_url"):
            data["original_url"] = raw_text if raw_text.startswith("http") else url
        seq = _append_to_excel(data)
        data["seq"] = seq
    return jsonify(result)


@app.route("/data")
def get_data():
    return jsonify(_read_all_from_excel())


@app.route("/audit")
def audit():
    """扫描 Excel，返回所有脏行的序号和当前内容"""
    rows = _read_all_from_excel()
    dirty = [
        {
            "seq":     row.get("序号"),
            "title":   row.get("标题", ""),
            "url":     str(row.get("原链接") or row.get("笔记链接") or ""),
            "content": str(row.get("正文内容") or ""),
        }
        for row in rows if _is_dirty(row)
    ]
    return jsonify({"total": len(rows), "dirty": dirty})


@app.route("/audit/fix", methods=["POST"])
def audit_fix():
    """
    对所有脏行重新采集并原地更新 Excel。
    使用 Server-Sent Events 实时推送每行进度。
    """
    rows = _read_all_from_excel()
    dirty = [r for r in rows if _is_dirty(r)]

    def generate():
        total = len(dirty)
        if total == 0:
            yield f"data: {json.dumps({'type':'done','fixed':0,'total':0})}\n\n"
            return

        fixed = 0
        for i, row in enumerate(dirty, 1):
            seq = row.get("序号")
            url = str(row.get("原链接") or row.get("笔记链接") or "").strip()

            yield f"data: {json.dumps({'type':'progress','seq':seq,'idx':i,'total':total,'status':'scraping'})}\n\n"

            result = scrape_note(url)
            if result["success"]:
                _update_excel_row(seq, result["data"])
                fixed += 1
                yield f"data: {json.dumps({'type':'progress','seq':seq,'idx':i,'total':total,'status':'ok','title':result['data'].get('title','')})}\n\n"
            else:
                yield f"data: {json.dumps({'type':'progress','seq':seq,'idx':i,'total':total,'status':'fail','error':result.get('error','')})}\n\n"

        yield f"data: {json.dumps({'type':'done','fixed':fixed,'total':total})}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/export")
def export():
    _ensure_excel()
    return send_file(
        EXCEL_PATH,
        as_attachment=True,
        download_name="xhs_notes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/clear", methods=["POST"])
def clear_data():
    """清空 Excel 数据（保留表头），供下一轮采集使用"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "小红书笔记"
    ws.append(HEADERS)
    _style_header(ws)
    wb.save(EXCEL_PATH)
    return jsonify({"ok": True, "message": "数据已清空"})


if __name__ == "__main__":
    _ensure_excel()
    print("\n🌸 小红书笔记采集器已启动")
    print("👉 请在浏览器中打开: http://localhost:5001\n")
    app.run(host="0.0.0.0", port=5001, debug=False)
